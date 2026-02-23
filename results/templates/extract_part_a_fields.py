#!/usr/bin/env python3
"""
Extract all field descriptions from Part-A-Template.xlsx and produce a
structured reference document (part-a-fields.txt) that can be used to
programmatically fill the template later.

Uses openpyxl for spreadsheet reading and zipfile/xml for data-validation
extraction (openpyxl drops the x14-namespace validations with a warning).
"""

import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import re
from pathlib import Path

XLSX = Path(__file__).parent / "Part-A-Template.xlsx"
OUTPUT = Path(__file__).parent / "part-a-fields.txt"

# -- helpers ----------------------------------------------------------------

def _expand_sqref(sqref_str):
    """Expand an sqref like 'B16:B17 B22:B26 B20 B12:B14' into individual cell refs."""
    from openpyxl.utils import range_boundaries, get_column_letter
    cells = []
    for part in sqref_str.strip().split():
        if ":" in part:
            min_col, min_row, max_col, max_row = range_boundaries(part)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cells.append(f"{get_column_letter(c)}{r}")
        else:
            cells.append(part)
    return cells


def _read_dropdown_lists(wb):
    """Read the 'lists' sheet and return named ranges of values."""
    ws = wb["lists"]
    result = {}
    vals_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    result["lists!$A$1:$A$2"] = vals_a
    vals_c = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=3).value]
    result["lists!$C$1:$C$5"] = vals_c
    vals_e = [ws.cell(row=r, column=5).value for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=5).value]
    result["lists!$E$1:$E$5"] = vals_e
    vals_g = [ws.cell(row=r, column=7).value for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=7).value]
    result["lists!$G$1"] = [vals_g[0]] if len(vals_g) >= 1 else []
    result["lists!$G$2"] = [vals_g[1]] if len(vals_g) >= 2 else []
    return result


def _extract_data_validations(xlsx_path):
    """
    Parse data validations from the raw XML inside the .xlsx zip.
    Returns {sheet_index_1based: [{sqref, type, prompt, formula_ref}, ...]}.
    """
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_x14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    ns_xm = "http://schemas.microsoft.com/office/excel/2006/main"

    validations = {}

    with zipfile.ZipFile(xlsx_path) as z:
        for fname in sorted(z.namelist()):
            m = re.match(r"xl/worksheets/sheet(\d+)\.xml", fname)
            if not m:
                continue
            sheet_idx = int(m.group(1))
            tree = ET.parse(z.open(fname))
            root = tree.getroot()
            sheet_dvs = []

            # Standard dataValidation elements
            for dv in root.iter(f"{{{ns_main}}}dataValidation"):
                sqref = dv.get("sqref", "")
                vtype = dv.get("type", "")
                prompt = dv.get("prompt", "")
                f1_el = dv.find(f"{{{ns_main}}}formula1")
                f1 = f1_el.text if f1_el is not None else ""
                sheet_dvs.append({"sqref": sqref, "type": vtype, "prompt": prompt, "formula_ref": f1})

            # x14:dataValidation elements (extended, inside extLst)
            for dv in root.iter(f"{{{ns_x14}}}dataValidation"):
                vtype = dv.get("type", "")
                prompt = dv.get("prompt", "")
                sqref_el = dv.find(f"{{{ns_xm}}}sqref")
                sqref = sqref_el.text if sqref_el is not None else ""
                f1_el = dv.find(f"{{{ns_x14}}}formula1")
                f1 = ""
                if f1_el is not None:
                    xm_f = f1_el.find(f"{{{ns_xm}}}f")
                    if xm_f is not None:
                        f1 = xm_f.text or ""
                sheet_dvs.append({"sqref": sqref, "type": vtype, "prompt": prompt, "formula_ref": f1})

            if sheet_dvs:
                validations[sheet_idx] = sheet_dvs

    return validations


# -- main -------------------------------------------------------------------

def main():
    wb = openpyxl.load_workbook(str(XLSX))
    dropdown_lists = _read_dropdown_lists(wb)
    raw_dvs = _extract_data_validations(str(XLSX))

    # Map sheet index -> sheet name (1-based, matching XML naming)
    sheet_names_ordered = wb.sheetnames
    idx_to_name = {i + 1: name for i, name in enumerate(sheet_names_ordered)}

    # Build per-sheet, per-cell validation lookup
    cell_validations = {}
    for idx, dvs in raw_dvs.items():
        sname = idx_to_name.get(idx, f"Sheet{idx}")
        if sname not in cell_validations:
            cell_validations[sname] = {}
        for dv in dvs:
            options = dropdown_lists.get(dv["formula_ref"], [])
            for cell in _expand_sqref(dv["sqref"]):
                cell_validations[sname][cell] = {
                    "type": dv["type"],
                    "prompt": dv["prompt"],
                    "formula_ref": dv["formula_ref"],
                    "options": options,
                }

    lines = []
    sep = "=" * 80

    lines.append(sep)
    lines.append("PART A TEMPLATE -- FIELD REFERENCE")
    lines.append("Source: Part-A-Template.xlsx")
    lines.append("Generated by: extract_part_a_fields.py")
    lines.append(sep)
    lines.append("")
    lines.append("IMPORTANT: Do NOT rename, delete or add sheets, rows or columns.")
    lines.append("")

    # Process each sheet except 'lists'
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == "lists":
            continue

        ws = wb[sheet_name]
        lines.append(sep)
        lines.append(f"SHEET: \"{sheet_name.strip()}\"")
        lines.append(f"  Rows: {ws.max_row}  |  Columns: {ws.max_column}")
        lines.append(sep)
        lines.append("")

        cv = cell_validations.get(sheet_name, {})

        for row_idx in range(1, ws.max_row + 1):
            row_cells = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_cells.append(cell)

            if not row_cells:
                continue

            # Header/instruction row
            if row_idx == 1 and any("DO NOT RENAME" in str(c.value) for c in row_cells):
                lines.append(f"  Row {row_idx}: [INSTRUCTION] {row_cells[0].value}")
                lines.append("")
                continue

            for cell in row_cells:
                coord = cell.coordinate
                val = cell.value
                col_letter = re.match(r"([A-Z]+)", coord).group(1)

                is_label = col_letter == "A"
                is_value_cell = col_letter in ("B", "C", "D", "E", "F", "G", "H")

                val_info = cv.get(coord)

                if is_label:
                    label_text = str(val).strip()
                    b_cell = ws.cell(row=row_idx, column=2)
                    c_cell = ws.cell(row=row_idx, column=3)
                    b_val = b_cell.value
                    c_val = c_cell.value
                    b_coord = b_cell.coordinate
                    c_coord = c_cell.coordinate
                    b_info = cv.get(b_coord)
                    c_info = cv.get(c_coord)

                    lines.append(f"  [{coord}] FIELD: {label_text}")

                    if b_val is not None:
                        if isinstance(b_val, str) and not b_val.startswith("="):
                            lines.append(f"       [{b_coord}] Placeholder/Hint: \"{b_val}\"")
                        elif isinstance(b_val, (int, float)):
                            lines.append(f"       [{b_coord}] Pre-filled value: {b_val}")
                        elif isinstance(b_val, str) and b_val.startswith("="):
                            lines.append(f"       [{b_coord}] Formula: {b_val}")

                    if b_info:
                        dtype = b_info["type"]
                        if dtype == "list":
                            opts = b_info["options"]
                            prompt = b_info.get("prompt", "")
                            lines.append(f"       [{b_coord}] DROPDOWN ({dtype}): {opts}")
                            if prompt:
                                lines.append(f"       Prompt: \"{prompt}\"")
                        elif dtype == "whole":
                            lines.append(f"       [{b_coord}] VALIDATION: whole number")
                            if b_info.get("prompt"):
                                lines.append(f"       Prompt: \"{b_info['prompt']}\"")

                    if c_val is not None:
                        lines.append(f"       [{c_coord}] Conditional note: \"{c_val}\"")
                    if c_info:
                        if c_info["type"] == "list":
                            lines.append(f"       [{c_coord}] DROPDOWN: {c_info['options']}")

                    lines.append("")

                elif is_value_cell and not is_label:
                    a_val = ws.cell(row=row_idx, column=1).value
                    if a_val is None:
                        if isinstance(val, str) and val.startswith("="):
                            lines.append(f"  [{coord}] Formula: {val}")
                        elif isinstance(val, str) and len(val) > 10:
                            lines.append(f"  [{coord}] Note: \"{val[:300]}\"")
                        elif isinstance(val, (int, float)):
                            lines.append(f"  [{coord}] Pre-filled: {val}")
                        if val_info and val_info["type"] == "list":
                            lines.append(f"       DROPDOWN: {val_info['options']}")
                        lines.append("")

        lines.append("")

    # -- Appendix: Dropdown option lists ------------------------------------
    lines.append(sep)
    lines.append("APPENDIX A: DROPDOWN OPTION LISTS (from 'lists' sheet)")
    lines.append(sep)
    lines.append("")
    lines.append("  Yes/No (lists!A1:A2):")
    lines.append("    - Yes")
    lines.append("    - No")
    lines.append("")
    lines.append("  Organisation Type (lists!C1:C5):")
    lines.append("    - SME   (independent enterprise, max 250 employees, max 50M EUR turnover)")
    lines.append("    - MID   (mid-cap, max 2000 employees, max 400M EUR turnover)")
    lines.append("    - IND   (large industrial enterprise)")
    lines.append("    - AC    (academia -- universities and research institutes)")
    lines.append("    - OTHER (governmental agencies, industry consortia, etc.)")
    lines.append("")
    lines.append("  Participant Role (lists!E1:E5):")
    lines.append("    - Application expert")
    lines.append("    - HPC expert")
    lines.append("    - Software provider")
    lines.append("    - Software developer")
    lines.append("    - HPC centre")
    lines.append("")
    lines.append("  Participant Type (lists!G1:G2):")
    lines.append("    - MP  (Main Participant -- always Participant 1)")
    lines.append("    - SP  (Supporting Participant -- Participants 2 and 3)")
    lines.append("")

    # -- Appendix: Requested Funding formulas -------------------------------
    lines.append(sep)
    lines.append("APPENDIX B: REQUESTED FUNDING -- FORMULAS (auto-calculated, do NOT overwrite)")
    lines.append(sep)
    lines.append("")
    lines.append("  G4 = SUM(E4:F4)   -- Total costs for Participant 1")
    lines.append("  G5 = SUM(E5:F5)   -- Total costs for Participant 2")
    lines.append("  G6 = SUM(E6:F6)   -- Total costs for Participant 3")
    lines.append("  D7 = SUM(D4:D6)   -- Sum of Effort (PM)")
    lines.append("  E7 = SUM(E4:E6)   -- Sum of Personnel Costs")
    lines.append("  F7 = SUM(F4:F6)   -- Sum of Other Direct Costs")
    lines.append("  G7 = SUM(G4:G6)   -- Sum of Total Costs")
    lines.append("  H7 = SUM(H4:H6)   -- Sum of Requested Funding")
    lines.append("")

    # -- Appendix: Quick fill map -------------------------------------------
    lines.append(sep)
    lines.append("APPENDIX C: QUICK-FILL CELL MAP (every cell that needs user input)")
    lines.append(sep)
    lines.append("")
    lines.append("  Sheet: \"General Info\"")
    lines.append("    B2   - Innovation Study title (text)")
    lines.append("    B3   - Innovation Study acronym (text)")
    lines.append("    B4   - Coordinating person name (text)")
    lines.append("    B5   - Coordinator email (text)")
    lines.append("    B6   - Keywords, up to 5 -- include AI solution type and industry domain (text)")
    lines.append("    B7   - Software solutions, e.g. TensorFlow, PyTorch (text)")
    lines.append("    B8   - Data access plan (text)")
    lines.append("    B9   - HPC resources, mention EuroHPC JU access if applicable (text)")
    lines.append("    B10  - Number of participating organisations (integer 1-3, validated)")
    lines.append("    B11  - Abstract, up to 100 words (text)")
    lines.append("    B12  - Extension of OC1 Type-2 study? (Yes/No dropdown)")
    lines.append("    C12  - If yes: study number and name from Open Call-1 Type-2 (text)")
    lines.append("    B13  - Discussed with NCC? (Yes/No dropdown)")
    lines.append("    C13  - If yes: which NCC? (text)")
    lines.append("    B14  - Discussed with AI Factory or AI Factory Antenna? (Yes/No dropdown)")
    lines.append("    C14  - If yes: which one? (text)")
    lines.append("    B16  - Similar proposal already submitted in FFplus Call1? (Yes/No dropdown)")
    lines.append("    B17  - Similar activities funded through past Fortissimo/EC/national projects? (Yes/No dropdown)")
    lines.append("    B18  - If yes: list previous projects (text)")
    lines.append("    B20  - Willing to respond to FFplus surveys? (Yes/No dropdown)")
    lines.append("    B22  - Became aware via FFplus Website? (Yes/No dropdown)")
    lines.append("    C22  - If yes: details (text)")
    lines.append("    B23  - Became aware via Electronic Media? (Yes/No dropdown)")
    lines.append("    C23  - If yes: details (text)")
    lines.append("    B24  - Became aware via Webinars or presentation? (Yes/No dropdown)")
    lines.append("    C24  - If yes: details (text)")
    lines.append("    B25  - Became aware via Other media? (Yes/No dropdown)")
    lines.append("    C25  - If yes: details (text)")
    lines.append("    B26  - Became aware via Personal Contacts? (Yes/No dropdown)")
    lines.append("    C26  - If yes: details (text)")
    lines.append("    B27  - Became aware via Other? (Yes/No dropdown)")
    lines.append("    C27  - If yes: details (text)")
    lines.append("")
    lines.append("  Sheet: \"Participant 1\"")
    lines.append("    B2   - Participant Org Number (pre-filled: 1, do not change)")
    lines.append("    B3   - Participant Type (pre-filled/locked: MP, do not change)")
    lines.append("    B4   - Company legal name (text)")
    lines.append("    B5   - Short name (text, optional convenience)")
    lines.append("    B6   - Full legal address (text)")
    lines.append("    B7   - Country, 2-letter ISO code (text)")
    lines.append("    B8   - Contact person name (text)")
    lines.append("    B9   - Contact email (text)")
    lines.append("    B10  - Telephone number (text)")
    lines.append("    B11  - Legal signatory name (text)")
    lines.append("    B12  - Legal signatory email (text)")
    lines.append("    B13  - PIC number from EC Participant Register (text)")
    lines.append("    B14  - Organisation type (dropdown: SME / MID / IND / AC / OTHER)")
    lines.append("    B16  - First EU Project? (Yes/No dropdown)")
    lines.append("")
    lines.append("  Sheet: \"Participant 2\" (only if >= 2 participants)")
    lines.append("    B2   - Participant Org Number (pre-filled: 2, do not change)")
    lines.append("    B3   - Participant Type (pre-filled/locked: SP, do not change)")
    lines.append("    B4   - Company legal name (text)")
    lines.append("    B5   - Short name (text)")
    lines.append("    B6   - Full legal address (text)")
    lines.append("    B7   - Country, 2-letter ISO code (text)")
    lines.append("    B8   - Contact person name (text)")
    lines.append("    B9   - Contact email (text)")
    lines.append("    B10  - Telephone number (text)")
    lines.append("    B11  - Legal signatory name (text)")
    lines.append("    B12  - Legal signatory email (text)")
    lines.append("    B13  - PIC number (text)")
    lines.append("    B14  - Organisation type (dropdown: SME / MID / IND / AC / OTHER)")
    lines.append("    B16  - Role (dropdown: Application expert / HPC expert / Software provider / Software developer / HPC centre)")
    lines.append("    B17  - First EU Project? (Yes/No dropdown)")
    lines.append("")
    lines.append("  Sheet: \"Participant 3\" (only if 3 participants)")
    lines.append("    B2   - Participant Org Number (pre-filled: 3, do not change)")
    lines.append("    B3   - Participant Type (pre-filled/locked: SP, do not change)")
    lines.append("    B4   - Company legal name (text)")
    lines.append("    B5   - Short name (text)")
    lines.append("    B6   - Full legal address (text)")
    lines.append("    B7   - Country, 2-letter ISO code (text)")
    lines.append("    B8   - Contact person name (text)")
    lines.append("    B9   - Contact email (text)")
    lines.append("    B10  - Telephone number (text)")
    lines.append("    B11  - Legal signatory name (text)")
    lines.append("    B12  - Legal signatory email (text)")
    lines.append("    B13  - PIC number (text)")
    lines.append("    B14  - Organisation type (dropdown: SME / MID / IND / AC / OTHER)")
    lines.append("    B16  - Role (dropdown: Application expert / HPC expert / Software provider / Software developer / HPC centre)")
    lines.append("    B17  - First EU Project? (Yes/No dropdown)")
    lines.append("")
    lines.append("  Sheet: \"Requested Funding\"")
    lines.append("    B4   - Participant 1 short name (text)")
    lines.append("    C4   - Funding rate for Participant 1 (pre-filled: 1 = 100%)")
    lines.append("    D4   - Effort in person-months, Participant 1 (number)")
    lines.append("    E4   - Personnel Costs in EUR, Participant 1 (number)")
    lines.append("    F4   - Other Direct Costs in EUR, Participant 1 (number)")
    lines.append("    G4   - [FORMULA] Total costs Participant 1 = SUM(E4:F4)")
    lines.append("    H4   - Requested Funding in EUR, Participant 1 (number)")
    lines.append("    B5   - Participant 2 short name (text)")
    lines.append("    C5   - Funding rate for Participant 2 (number, 0-1)")
    lines.append("    D5   - Effort in person-months, Participant 2 (number)")
    lines.append("    E5   - Personnel Costs in EUR, Participant 2 (number)")
    lines.append("    F5   - Other Direct Costs in EUR, Participant 2 (number)")
    lines.append("    G5   - [FORMULA] Total costs Participant 2 = SUM(E5:F5)")
    lines.append("    H5   - Requested Funding in EUR, Participant 2 (number)")
    lines.append("    B6   - Participant 3 short name (text)")
    lines.append("    C6   - Funding rate for Participant 3 (number, 0-1)")
    lines.append("    D6   - Effort in person-months, Participant 3 (number)")
    lines.append("    E6   - Personnel Costs in EUR, Participant 3 (number)")
    lines.append("    F6   - Other Direct Costs in EUR, Participant 3 (number)")
    lines.append("    G6   - [FORMULA] Total costs Participant 3 = SUM(E6:F6)")
    lines.append("    H6   - Requested Funding in EUR, Participant 3 (number)")
    lines.append("    Row 7 - [ALL FORMULAS] Sum row, auto-calculated")
    lines.append("")
    lines.append(sep)
    lines.append("END OF REFERENCE DOCUMENT")
    lines.append(sep)

    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Written {len(lines)} lines to {OUTPUT}")
    print("Done.")


if __name__ == "__main__":
    main()
